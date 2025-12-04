from typing import Dict

import openpyxl
book = openpyxl.load_workbook("C:\\Users\\Hp\\PycharmProjects\\PythonTesting\\selenium\\Book1.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1,column=2)
# print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['a5'].value)

for r in range(1,sheet.max_row+1):
    if sheet.cell(row=r, column=1).value == "TestCase2":
        for c in range(1,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value

print(Dict)